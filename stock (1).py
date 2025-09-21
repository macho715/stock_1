import pandas as pd
import numpy as np
from datetime import datetime, date
from collections import defaultdict, Counter
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import warnings
warnings.filterwarnings('ignore')

class InventoryTracker:
    def __init__(self, excel_file_path):
        """
        창고 재고 추적 시스템 초기화
        
        Args:
            excel_file_path (str): Excel 파일 경로
        """
        self.excel_file = excel_file_path
        self.workbook = None
        self.case_data = defaultdict(list)  # CASE -> [(date, location), ...]
        self.out_data = defaultdict(list)   # 출고 데이터
        self.global_max_date = None
        
        # 컬럼 매핑 (VBA와 동일: B=2, D=4, H=8 -> 0-based index)
        self.CASE_COL = 1    # Column B (0-based: 1)
        self.LOCATION_COL = 3  # Column D (0-based: 3)  
        self.DATE_COL = 7    # Column H (0-based: 7)
    
    def load_workbook(self):
        """Excel 워크북 로드"""
        try:
            self.workbook = pd.ExcelFile(self.excel_file)
            print(f"✅ 워크북 로드 완료: {len(self.workbook.sheet_names)}개 시트 발견")
            return True
        except Exception as e:
            print(f"❌ 워크북 로드 실패: {e}")
            return False
    
    def normalize_date(self, date_val, current_year=2024):
        """날짜 정규화: yyyy-mm-dd 형식으로 변환"""
        if pd.isna(date_val) or date_val == "":
            return ""
        
        try:
            if isinstance(date_val, (datetime, date)):
                return date_val.strftime('%Y-%m-%d')
            elif isinstance(date_val, str):
                date_val = date_val.strip()
                if date_val == "":
                    return ""
                
                # "16-Feb", "23-Apr" 형식 처리 (년도 없는 경우)
                if re.match(r'^\d{1,2}-[A-Za-z]{3}$', date_val):
                    try:
                        # 현재 년도 추가하여 파싱
                        full_date = f"{date_val}-{current_year}"
                        parsed_date = pd.to_datetime(full_date, format='%d-%b-%Y')
                        return parsed_date.strftime('%Y-%m-%d')
                    except:
                        pass
                
                # 일반적인 날짜 형식 처리
                parsed_date = pd.to_datetime(date_val)
                return parsed_date.strftime('%Y-%m-%d')
            else:
                parsed_date = pd.to_datetime(str(date_val))
                return parsed_date.strftime('%Y-%m-%d')
        except:
            return ""
    
    def is_out_sheet(self, sheet_name):
        """출고 시트 판별: 이름에 OUT 또는 DISPATCH 포함"""
        sheet_upper = sheet_name.upper()
        return 'OUT' in sheet_upper or 'DISPATCH' in sheet_upper
    
    def process_sheet(self, sheet_name):
        """개별 시트 처리 - 통합분석 파일 구조에 맞게 수정"""
        try:
            # 시트 데이터 로드 (헤더 1행, 데이터 2행부터)
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name, header=0)
            
            print(f"📋 {sheet_name} 처리 중... (행: {len(df)}, 열: {len(df.columns)})")
            print(f"   컬럼: {list(df.columns)}")
            
            # 시트별로 다른 처리 로직 적용
            if sheet_name == "종합_SKU요약":
                processed_count = self.process_sku_summary_sheet(df, sheet_name)
            elif sheet_name == "날짜별_추이":
                processed_count = self.process_date_trend_sheet(df, sheet_name)
            elif sheet_name == "월별_분석":
                processed_count = self.process_monthly_analysis_sheet(df, sheet_name)
            elif sheet_name == "창고별_현황":
                processed_count = self.process_warehouse_status_sheet(df, sheet_name)
            elif sheet_name == "분석_통계":
                processed_count = self.process_statistics_sheet(df, sheet_name)
            else:
                # 기존 로직 유지 (일반 재고 시트용)
                processed_count = self.process_general_sheet(df, sheet_name)
            
            sheet_type = "출고" if self.is_out_sheet(sheet_name) else "입고"
            print(f"   ✅ {processed_count}건 처리 완료 ({sheet_type})")
            
        except Exception as e:
            print(f"❌ {sheet_name} 처리 실패: {e}")
            import traceback
            print(f"   상세 오류: {traceback.format_exc()}")
    
    def process_sku_summary_sheet(self, df, sheet_name):
        """종합_SKU요약 시트 처리"""
        processed_count = 0
        
        for idx, row in df.iterrows():
            if pd.notna(row['SKU']):
                case_no = str(row['SKU']).strip()
                location = str(row['Last_Location']).strip() if pd.notna(row['Last_Location']) else ""
                date_raw = row['Last_Seen'] if pd.notna(row['Last_Seen']) else ""
                date_normalized = self.normalize_date(date_raw)
                
                if case_no and case_no.lower() != 'nan':
                    entry = (date_normalized, location)
                    
                    # 상태에 따라 입고/출고 분류
                    if 'OUT' in str(row.get('Status', '')):
                        if entry not in self.out_data[case_no]:
                            self.out_data[case_no].append(entry)
                            processed_count += 1
                    else:
                        if entry not in self.case_data[case_no]:
                            self.case_data[case_no].append(entry)
                            processed_count += 1
        
        return processed_count
    
    def process_date_trend_sheet(self, df, sheet_name):
        """날짜별_추이 시트 처리"""
        processed_count = 0
        
        for idx, row in df.iterrows():
            if pd.notna(row['Date']):
                case_no = f"날짜추이_{row['Date'].strftime('%Y-%m-%d')}"
                location = f"SKU수량:{row['SKU_Count']}"
                date_raw = row['Date']
                date_normalized = self.normalize_date(date_raw)
                
                entry = (date_normalized, location)
                if entry not in self.case_data[case_no]:
                    self.case_data[case_no].append(entry)
                    processed_count += 1
        
        return processed_count
    
    def process_monthly_analysis_sheet(self, df, sheet_name):
        """월별_분석 시트 처리"""
        processed_count = 0
        
        for idx, row in df.iterrows():
            if pd.notna(row['Month_Key']):
                case_no = str(row['Month_Key']).strip()
                location = f"IN:{row['Total_IN']}_OUT:{row['Total_OUT']}"
                # 월 키를 날짜로 변환 (예: 2024-06 -> 2024-06-01)
                date_raw = f"{row['Month_Key']}-01"
                date_normalized = self.normalize_date(date_raw)
                
                entry = (date_normalized, location)
                if entry not in self.case_data[case_no]:
                    self.case_data[case_no].append(entry)
                    processed_count += 1
        
        return processed_count
    
    def process_warehouse_status_sheet(self, df, sheet_name):
        """창고별_현황 시트 처리"""
        processed_count = 0
        
        for idx, row in df.iterrows():
            if pd.notna(row['Warehouse']):
                case_no = f"창고_{row['Warehouse']}"
                location = f"현재재고:{row['Current_Stock']}_총이력:{row['Total_Historical']}"
                date_normalized = datetime.now().strftime('%Y-%m-%d')  # 현재 날짜 사용
                
                entry = (date_normalized, location)
                if entry not in self.case_data[case_no]:
                    self.case_data[case_no].append(entry)
                    processed_count += 1
        
        return processed_count
    
    def process_statistics_sheet(self, df, sheet_name):
        """분석_통계 시트 처리"""
        processed_count = 0
        
        for idx, row in df.iterrows():
            if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                key = str(row.iloc[0]).strip()
                value = str(row.iloc[1]).strip()
                
                if key and key != 'nan' and value and value != 'nan':
                    case_no = f"통계_{key}"
                    location = value
                    date_normalized = datetime.now().strftime('%Y-%m-%d')  # 현재 날짜 사용
                    
                    entry = (date_normalized, location)
                    if entry not in self.case_data[case_no]:
                        self.case_data[case_no].append(entry)
                        processed_count += 1
        
        return processed_count
    
    def process_general_sheet(self, df, sheet_name):
        """일반 재고 시트 처리 (기존 로직)"""
        processed_count = 0
        
        if len(df.columns) <= max(self.CASE_COL, self.LOCATION_COL, self.DATE_COL):
            print(f"⚠️  {sheet_name}: 컬럼 수 부족 (필요: {max(self.CASE_COL, self.LOCATION_COL, self.DATE_COL)+1}, 실제: {len(df.columns)}), 건너뜀")
            return 0
        
        # 필요한 컬럼만 추출 (안전하게)
        case_col = df.iloc[:, self.CASE_COL] if len(df.columns) > self.CASE_COL else pd.Series()
        location_col = df.iloc[:, self.LOCATION_COL] if len(df.columns) > self.LOCATION_COL else pd.Series()
        date_col = df.iloc[:, self.DATE_COL] if len(df.columns) > self.DATE_COL else pd.Series()
        
        # 빈 CASE_NO 제거
        valid_rows = case_col.notna() & (case_col.astype(str).str.strip() != '')
        
        is_out = self.is_out_sheet(sheet_name)
        
        for idx in df.index[valid_rows]:
            case_no = str(case_col.iloc[idx]).strip()
            location = str(location_col.iloc[idx]).strip() if idx < len(location_col) and pd.notna(location_col.iloc[idx]) else ""
            date_raw = date_col.iloc[idx] if idx < len(date_col) else ""
            date_normalized = self.normalize_date(date_raw)
            
            if case_no and case_no.lower() != 'nan':
                entry = (date_normalized, location)
                
                if is_out:
                    # 출고 데이터에 중복 체크 후 추가
                    if entry not in self.out_data[case_no]:
                        self.out_data[case_no].append(entry)
                        processed_count += 1
                else:
                    # 입고 데이터에 중복 체크 후 추가
                    if entry not in self.case_data[case_no]:
                        self.case_data[case_no].append(entry)
                        processed_count += 1
        
        return processed_count
    
    def calculate_global_max_date(self):
        """전역 최신 날짜 계산 (입고 데이터 기준)"""
        max_date = None
        
        for case_entries in self.case_data.values():
            for date_str, _ in case_entries:
                try:
                    current_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    if max_date is None or current_date > max_date:
                        max_date = current_date
                except:
                    continue
        
        self.global_max_date = max_date
        if max_date:
            print(f"🗓️  전역 최신 스냅샷 날짜: {max_date}")
    
    def determine_status(self, case_no, in_entries, out_entries):
        """재고 상태 결정"""
        if out_entries:
            # 출고 기록이 있으면 OUT
            out_dates = [entry[0] for entry in out_entries if entry[0]]
            return "OUT (by OUTsheet)", f"OutDates: {', '.join(out_dates)}"
        
        elif in_entries:
            # 입고 기록만 있는 경우 - 최신 스냅샷과 비교
            sorted_entries = sorted(in_entries, key=lambda x: x[0])
            last_seen_str = sorted_entries[-1][0]
            
            try:
                last_seen_date = datetime.strptime(last_seen_str, '%Y-%m-%d').date()
                
                if self.global_max_date and last_seen_date < self.global_max_date:
                    return "OUT (absent in latest snapshot)", \
                           f"LastSeen: {last_seen_str} ; GlobalLatest: {self.global_max_date}"
                else:
                    return "IN", ""
            except:
                return "IN", ""
        
        else:
            return "UNKNOWN", ""
    
    def create_summary(self):
        """요약 리포트 생성"""
        # 모든 CASE 수집
        all_cases = set(self.case_data.keys()) | set(self.out_data.keys())
        
        summary_data = []
        
        for case_no in sorted(all_cases):
            in_entries = sorted(self.case_data.get(case_no, []), key=lambda x: x[0])
            out_entries = sorted(self.out_data.get(case_no, []), key=lambda x: x[0])
            
            # 기본 정보
            count = len(in_entries)
            first_in_date = in_entries[0][0] if in_entries else ""
            all_in_dates = ", ".join([entry[0] for entry in in_entries])
            last_seen = in_entries[-1][0] if in_entries else ""
            last_location = in_entries[-1][1] if in_entries else ""
            
            # 상태 결정
            status, note = self.determine_status(case_no, in_entries, out_entries)
            
            summary_data.append({
                'CASE_NO': case_no,
                'Count': count,
                'First_IN_Date': first_in_date,
                'All_IN_Dates': all_in_dates,
                'LastSeen': last_seen,
                'Last_Location': last_location,
                'Status': status,
                'Note': note
            })
        
        return pd.DataFrame(summary_data)
    
    def save_summary_to_excel(self, output_file=None):
        """요약 결과를 Excel 파일로 저장"""
        if output_file is None:
            output_file = self.excel_file.replace('.xlsx', '_with_summary.xlsx')
        
        try:
            # 기존 워크북 복사
            wb = openpyxl.load_workbook(self.excel_file)
            
            # 기존 Onhand_Summary 시트 삭제 (있다면)
            if 'Onhand_Summary' in wb.sheetnames:
                wb.remove(wb['Onhand_Summary'])
            
            # 새 요약 시트 생성
            summary_df = self.create_summary()
            ws = wb.create_sheet('Onhand_Summary')
            
            # 데이터프레임을 시트에 쓰기
            for r in dataframe_to_rows(summary_df, index=False, header=True):
                ws.append(r)
            
            # 컬럼 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_file)
            print(f"✅ 요약 파일 저장 완료: {output_file}")
            print(f"📊 총 {len(summary_df)}건 처리")
            return output_file
            
        except Exception as e:
            print(f"❌ 파일 저장 실패: {e}")
            return None
    
    def run_analysis(self):
        """전체 분석 실행"""
        print("🚀 재고 추적 분석 시작...")
        
        # 1. 워크북 로드
        if not self.load_workbook():
            return None
        
        # 2. 모든 시트 처리
        print("\n📂 시트 처리 중...")
        for sheet_name in self.workbook.sheet_names:
            if sheet_name not in ['Onhand_Summary']:  # 요약 시트 제외
                self.process_sheet(sheet_name)
        
        # 3. 전역 최신 날짜 계산
        print("\n📅 날짜 분석 중...")
        self.calculate_global_max_date()
        
        # 4. 요약 리포트 생성 및 저장
        print("\n📋 요약 리포트 생성 중...")
        return self.save_summary_to_excel()
    
    def get_status_summary(self):
        """상태별 요약 통계"""
        summary_df = self.create_summary()
        status_counts = summary_df['Status'].value_counts()
        
        print("\n📈 상태별 통계:")
        for status, count in status_counts.items():
            print(f"   {status}: {count}건")
        
        return status_counts


# 사용 예시 및 메인 실행 함수
def main(excel_file_path, output_file=None):
    """
    메인 실행 함수
    
    Args:
        excel_file_path (str): 분석할 Excel 파일 경로
        output_file (str): 출력 파일 경로 (선택사항)
    
    Returns:
        str: 생성된 요약 파일 경로
    """
    print("=" * 60)
    print("🏭 HVDC 프로젝트 재고 추적 시스템 v2.0")
    print("=" * 60)
    
    tracker = InventoryTracker(excel_file_path)
    
    # 분석 실행
    result_file = tracker.run_analysis()
    
    if result_file:
        # 상태별 통계 출력
        status_counts = tracker.get_status_summary()
        
        # 요약 데이터프레임 미리보기
        summary_df = tracker.create_summary()
        print(f"\n📋 요약 데이터 미리보기 (상위 10건):")
        print(summary_df.head(10).to_string(index=False))
        
        # 중요 통계
        print(f"\n📊 핵심 통계:")
        print(f"   📦 총 처리 CASE 수: {len(summary_df)}")
        print(f"   📥 총 입고 건수: {sum(len(entries) for entries in tracker.case_data.values())}")
        print(f"   📤 총 출고 건수: {sum(len(entries) for entries in tracker.out_data.values())}")
        print(f"   🔄 처리된 시트 수: {len(tracker.workbook.sheet_names) - 1}")  # 요약시트 제외
        
        # 출력 파일 이름 변경 (요청시)
        if output_file and output_file != result_file:
            try:
                import shutil
                shutil.copy2(result_file, output_file)
                print(f"   📁 최종 파일: {output_file}")
                return output_file
            except Exception as e:
                print(f"   ⚠️ 파일 복사 실패: {e}, 원본 사용: {result_file}")
        
        return result_file
    else:
        print("❌ 분석 실패")
        return None

def analyze_hvdc_inventory(file_path, show_details=True):
    """
    HVDC 프로젝트 전용 재고 분석 함수
    
    Args:
        file_path (str): Excel 파일 경로
        show_details (bool): 상세 정보 출력 여부
    """
    print("🔍 HVDC 재고 상세 분석 시작...")
    
    tracker = InventoryTracker(file_path)
    
    if not tracker.load_workbook():
        return None
    
    # 시트별 상세 분석
    sheet_analysis = {}
    for sheet_name in tracker.workbook.sheet_names:
        if sheet_name != 'Onhand_Summary':
            tracker.process_sheet(sheet_name)
            
            # 시트별 통계
            case_count = len([case for case, entries in tracker.case_data.items() if any(sheet_name in str(e) for e in entries)])
            sheet_analysis[sheet_name] = {
                'type': '출고' if tracker.is_out_sheet(sheet_name) else '입고',
                'case_count': case_count
            }
    
    if show_details:
        print(f"\n📊 시트별 분석 결과:")
        for sheet, info in sheet_analysis.items():
            print(f"   {sheet} ({info['type']}): {info['case_count']}개 CASE")
    
    tracker.calculate_global_max_date()
    summary_file = tracker.save_summary_to_excel()
    
    return summary_file

# 직접 실행 시 테스트
if __name__ == "__main__":
    # 실제 사용 예시
    print("🧪 테스트 모드")
    print("실제 사용법:")
    print("   result = main('your_inventory_file.xlsx')")
    print("   또는")  
    print("   result = analyze_hvdc_inventory('your_inventory_file.xlsx')")
    
    # 실제 파일이 있다면 주석 해제
    # test_file = "hvdc_inventory.xlsx"
    # result = main(test_file)
    
    # if result:
    #     print(f"\n🎉 분석 완료! 결과 파일: {result}")
    
    def is_out_sheet(self, sheet_name):
        """출고 시트 판별: 이름에 OUT 또는 DISPATCH 포함"""
        sheet_upper = sheet_name.upper()
        return 'OUT' in sheet_upper or 'DISPATCH' in sheet_upper
    
    def process_sheet(self, sheet_name):
        """개별 시트 처리"""
        try:
            # 시트 데이터 로드 (헤더 1행, 데이터 2행부터)
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name, header=0)
            
            if len(df.columns) <= max(self.CASE_COL, self.LOCATION_COL, self.DATE_COL):
                print(f"⚠️  {sheet_name}: 컬럼 수 부족, 건너뜀")
                return
            
            # 필요한 컬럼만 추출
            df_filtered = df.iloc[:, [self.CASE_COL, self.LOCATION_COL, self.DATE_COL]]
            df_filtered.columns = ['CASE_NO', 'LOCATION', 'DATE_RECEIVE']
            
            # 빈 CASE_NO 제거
            df_filtered = df_filtered[df_filtered['CASE_NO'].notna()]
            df_filtered = df_filtered[df_filtered['CASE_NO'].astype(str).str.strip() != '']
            
            is_out = self.is_out_sheet(sheet_name)
            processed_count = 0
            
            for _, row in df_filtered.iterrows():
                case_no = str(row['CASE_NO']).strip()
                location = str(row['LOCATION']).strip() if pd.notna(row['LOCATION']) else ""
                date_normalized = self.normalize_date(row['DATE_RECEIVE'])
                
                if case_no and date_normalized:
                    entry = (date_normalized, location)
                    
                    if is_out:
                        # 출고 데이터에 중복 체크 후 추가
                        if entry not in self.out_data[case_no]:
                            self.out_data[case_no].append(entry)
                            processed_count += 1
                    else:
                        # 입고 데이터에 중복 체크 후 추가
                        if entry not in self.case_data[case_no]:
                            self.case_data[case_no].append(entry)
                            processed_count += 1
            
            sheet_type = "출고" if is_out else "입고"
            print(f"📋 {sheet_name} ({sheet_type}): {processed_count}건 처리")
            
        except Exception as e:
            print(f"❌ {sheet_name} 처리 실패: {e}")
    
    def calculate_global_max_date(self):
        """전역 최신 날짜 계산 (입고 데이터 기준)"""
        max_date = None
        
        for case_entries in self.case_data.values():
            for date_str, _ in case_entries:
                try:
                    current_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    if max_date is None or current_date > max_date:
                        max_date = current_date
                except:
                    continue
        
        self.global_max_date = max_date
        if max_date:
            print(f"🗓️  전역 최신 스냅샷 날짜: {max_date}")
    
    def determine_status(self, case_no, in_entries, out_entries):
        """재고 상태 결정"""
        if out_entries:
            # 출고 기록이 있으면 OUT
            out_dates = [entry[0] for entry in out_entries if entry[0]]
            return "OUT (by OUTsheet)", f"OutDates: {', '.join(out_dates)}"
        
        elif in_entries:
            # 입고 기록만 있는 경우 - 최신 스냅샷과 비교
            sorted_entries = sorted(in_entries, key=lambda x: x[0])
            last_seen_str = sorted_entries[-1][0]
            
            try:
                last_seen_date = datetime.strptime(last_seen_str, '%Y-%m-%d').date()
                
                if self.global_max_date and last_seen_date < self.global_max_date:
                    return "OUT (absent in latest snapshot)", \
                           f"LastSeen: {last_seen_str} ; GlobalLatest: {self.global_max_date}"
                else:
                    return "IN", ""
            except:
                return "IN", ""
        
        else:
            return "UNKNOWN", ""
    
    def create_summary(self):
        """요약 리포트 생성"""
        # 모든 CASE 수집
        all_cases = set(self.case_data.keys()) | set(self.out_data.keys())
        
        summary_data = []
        
        for case_no in sorted(all_cases):
            in_entries = sorted(self.case_data.get(case_no, []), key=lambda x: x[0])
            out_entries = sorted(self.out_data.get(case_no, []), key=lambda x: x[0])
            
            # 기본 정보
            count = len(in_entries)
            first_in_date = in_entries[0][0] if in_entries else ""
            all_in_dates = ", ".join([entry[0] for entry in in_entries])
            last_seen = in_entries[-1][0] if in_entries else ""
            last_location = in_entries[-1][1] if in_entries else ""
            
            # 상태 결정
            status, note = self.determine_status(case_no, in_entries, out_entries)
            
            summary_data.append({
                'CASE_NO': case_no,
                'Count': count,
                'First_IN_Date': first_in_date,
                'All_IN_Dates': all_in_dates,
                'LastSeen': last_seen,
                'Last_Location': last_location,
                'Status': status,
                'Note': note
            })
        
        return pd.DataFrame(summary_data)
    
    def save_summary_to_excel(self, output_file=None):
        """요약 결과를 Excel 파일로 저장"""
        if output_file is None:
            output_file = self.excel_file.replace('.xlsx', '_with_summary.xlsx')
        
        try:
            # 기존 워크북 복사
            wb = openpyxl.load_workbook(self.excel_file)
            
            # 기존 Onhand_Summary 시트 삭제 (있다면)
            if 'Onhand_Summary' in wb.sheetnames:
                wb.remove(wb['Onhand_Summary'])
            
            # 새 요약 시트 생성
            summary_df = self.create_summary()
            ws = wb.create_sheet('Onhand_Summary')
            
            # 데이터프레임을 시트에 쓰기
            for r in dataframe_to_rows(summary_df, index=False, header=True):
                ws.append(r)
            
            # 컬럼 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_file)
            print(f"✅ 요약 파일 저장 완료: {output_file}")
            print(f"📊 총 {len(summary_df)}건 처리")
            return output_file
            
        except Exception as e:
            print(f"❌ 파일 저장 실패: {e}")
            return None
    
    def run_analysis(self):
        """전체 분석 실행"""
        print("🚀 재고 추적 분석 시작...")
        
        # 1. 워크북 로드
        if not self.load_workbook():
            return None
        
        # 2. 모든 시트 처리
        print("\n📂 시트 처리 중...")
        for sheet_name in self.workbook.sheet_names:
            if sheet_name not in ['Onhand_Summary']:  # 요약 시트 제외
                self.process_sheet(sheet_name)
        
        # 3. 전역 최신 날짜 계산
        print("\n📅 날짜 분석 중...")
        self.calculate_global_max_date()
        
        # 4. 요약 리포트 생성 및 저장
        print("\n📋 요약 리포트 생성 중...")
        return self.save_summary_to_excel()
    
    def get_status_summary(self):
        """상태별 요약 통계"""
        summary_df = self.create_summary()
        status_counts = summary_df['Status'].value_counts()
        
        print("\n📈 상태별 통계:")
        for status, count in status_counts.items():
            print(f"   {status}: {count}건")
        
        return status_counts


# 사용 예시 및 메인 실행 함수
def main(excel_file_path):
    """
    메인 실행 함수
    
    Args:
        excel_file_path (str): 분석할 Excel 파일 경로
    
    Returns:
        str: 생성된 요약 파일 경로
    """
    tracker = InventoryTracker(excel_file_path)
    
    # 분석 실행
    output_file = tracker.run_analysis()
    
    if output_file:
        # 상태별 통계 출력
        tracker.get_status_summary()
        
        # 요약 데이터프레임 미리보기
        summary_df = tracker.create_summary()
        print(f"\n📋 요약 데이터 미리보기 (상위 10건):")
        print(summary_df.head(10).to_string(index=False))
        
        return output_file
    else:
        print("❌ 분석 실패")
        return None

# 직접 실행 시 테스트
if __name__ == "__main__":
    # HVDC 프로젝트 재고 데이터 분석 - 통합분석 파일
    test_file = r"HVDC WH DATA\Stock On Hand Report_통합분석_20250919_1604.xlsx"
    
    print("=" * 60)
    print("🏭 HVDC 프로젝트 재고 추적 시스템 v2.0")
    print("=" * 60)
    
    result = main(test_file)
    
    if result:
        print(f"\n🎉 분석 완료! 결과 파일: {result}")
    else:
        print("❌ 분석 실패")