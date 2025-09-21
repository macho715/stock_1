#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HVDC SKU Master Hub - 최종 결과 Excel 생성기
각 Case No.(SKU)별로 개별 행으로 상세 정보를 표시하는 Excel 파일 생성
"""

import pandas as pd
import duckdb
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime

def load_sku_master_data():
    """SKU_MASTER 데이터 로드"""
    print("🔍 SKU_MASTER 데이터 로딩 중...")
    
    # Parquet 파일에서 직접 로드
    parquet_file = Path("out/SKU_MASTER.parquet")
    if parquet_file.exists():
        df = pd.read_parquet(parquet_file)
        print(f"📊 Parquet에서 {len(df)}개 레코드 로드 완료")
        return df
    
    # DuckDB에서 로드 (Parquet 파일이 없는 경우)
    duckdb_file = Path("out/sku_master.duckdb")
    if duckdb_file.exists():
        con = duckdb.connect(str(duckdb_file))
        df = con.execute("SELECT * FROM sku_master").df()
        con.close()
        print(f"🗃️ DuckDB에서 {len(df)}개 레코드 로드 완료")
        return df
    
    raise FileNotFoundError("SKU_MASTER 데이터 파일을 찾을 수 없습니다.")

def enhance_dataframe(df):
    """DataFrame 컬럼 정리 및 추가 정보 생성"""
    print("🔧 데이터 전처리 및 컬럼 정리...")
    
    # 컬럼명 정리 (한글 헤더 추가)
    column_mapping = {
        'SKU': 'SKU (Case No.)',
        'hvdc_code_norm': 'HVDC Code',
        'Vendor': 'Vendor',
        'Pkg': 'Package Count',
        'GW': 'Gross Weight (kg)',
        'CBM': 'Volume (m³)',
        'first_seen': 'First Seen Date',
        'last_seen': 'Last Seen Date',
        'Final_Location': 'Final Location',
        'FLOW_CODE': 'Flow Code',
        'flow_desc': 'Flow Description',
        'stock_qty': 'Current Stock',
        'sqm_cum': 'SQM Cumulative',
        'inv_match_status': 'Invoice Match Status',
        'err_gw': 'Weight Error (kg)',
        'err_cbm': 'Volume Error (m³)'
    }
    
    # 존재하는 컬럼만 매핑
    existing_columns = {k: v for k, v in column_mapping.items() if k in df.columns}
    df_enhanced = df.rename(columns=existing_columns)
    
    # Flow Code 설명 추가
    if 'Flow Code' in df_enhanced.columns:
        df_enhanced['Flow Type'] = df_enhanced['Flow Code'].map({
            0: '🚢 Pre Arrival',
            1: '🎯 Port → Site (직송)',
            2: '🏭 Port → WH → Site',
            3: '🏢 Port → WH → MOSB → Site',
            4: '🔄 Multi-hop'
        }).fillna('❓ Unknown')
    
    # 위치 분류 추가
    if 'Final Location' in df_enhanced.columns:
        def classify_location(location):
            if pd.isna(location):
                return '❓ Unknown'
            location_str = str(location).upper()
            if 'DSV' in location_str:
                return '📦 DSV Warehouse'
            elif 'MOSB' in location_str:
                return '🏢 MOSB'
            elif location_str in ['SHU', 'DAS', 'MIR', 'SITE']:
                return '🎯 Site Delivered'
            elif 'INDOOR' in location_str or 'OUTDOOR' in location_str:
                return '📦 Warehouse'
            else:
                return f'📍 {location}'
        
        df_enhanced['Location Type'] = df_enhanced['Final Location'].apply(classify_location)
    
    # 중량/부피 범위 분류
    if 'Gross Weight (kg)' in df_enhanced.columns:
        def classify_weight(weight):
            if pd.isna(weight):
                return '❓ Unknown'
            if weight < 1000:
                return '🪶 Light (<1톤)'
            elif weight < 5000:
                return '📦 Medium (1-5톤)'
            elif weight < 10000:
                return '🏗️ Heavy (5-10톤)'
            else:
                return '🚛 Very Heavy (>10톤)'
        
        df_enhanced['Weight Category'] = df_enhanced['Gross Weight (kg)'].apply(classify_weight)
    
    # 패키지 수량 범위 분류
    if 'Package Count' in df_enhanced.columns:
        def classify_packages(pkg):
            if pd.isna(pkg):
                return '❓ Unknown'
            if pkg == 1:
                return '📦 Single Package'
            elif pkg <= 5:
                return '📦📦 Small Batch (2-5)'
            elif pkg <= 20:
                return '📦📦📦 Medium Batch (6-20)'
            else:
                return '📦📦📦📦 Large Batch (>20)'
        
        df_enhanced['Package Category'] = df_enhanced['Package Count'].apply(classify_packages)
    
    # 데이터 품질 점수 계산
    quality_score = 0
    total_fields = 0
    
    key_fields = ['SKU (Case No.)', 'Vendor', 'Package Count', 'Gross Weight (kg)', 
                  'Volume (m³)', 'Final Location', 'Flow Code']
    
    for field in key_fields:
        if field in df_enhanced.columns:
            total_fields += 1
            quality_score += (~df_enhanced[field].isna()).astype(int)
    
    if total_fields > 0:
        df_enhanced['Data Quality Score'] = (quality_score / total_fields * 100).round(1)
        df_enhanced['Data Quality Level'] = df_enhanced['Data Quality Score'].apply(
            lambda x: '🟢 Excellent (≥90%)' if x >= 90 
            else '🟡 Good (70-89%)' if x >= 70 
            else '🟠 Fair (50-69%)' if x >= 50 
            else '🔴 Poor (<50%)'
        )
    
    print(f"✅ 데이터 전처리 완료 - {len(df_enhanced)}개 행, {len(df_enhanced.columns)}개 컬럼")
    return df_enhanced

def create_summary_stats(df):
    """요약 통계 생성"""
    print("📊 요약 통계 생성 중...")
    
    summary_data = {
        'Metric': [],
        'Value': [],
        'Description': []
    }
    
    # 기본 통계
    summary_data['Metric'].extend([
        'Total SKUs', 'Total Packages', 'Total Weight (tons)', 
        'Total Volume (m³)', 'Unique Vendors', 'Unique Locations'
    ])
    
    total_packages = df['Package Count'].sum() if 'Package Count' in df.columns else 0
    total_weight = df['Gross Weight (kg)'].sum() / 1000 if 'Gross Weight (kg)' in df.columns else 0
    total_volume = df['Volume (m³)'].sum() if 'Volume (m³)' in df.columns else 0
    unique_vendors = df['Vendor'].nunique() if 'Vendor' in df.columns else 0
    unique_locations = df['Final Location'].nunique() if 'Final Location' in df.columns else 0
    
    summary_data['Value'].extend([
        len(df), 
        int(total_packages), 
        f"{total_weight:,.1f}", 
        f"{total_volume:,.2f}",
        unique_vendors,
        unique_locations
    ])
    
    summary_data['Description'].extend([
        'SKU Master Hub에 통합된 총 Case Number 수',
        'Hitachi/Siemens 자재의 총 패키지 수량',
        '전체 자재의 총 중량 (톤)',
        '전체 자재의 총 부피 (세제곱미터)', 
        '프로젝트 참여 벤더 수 (Hitachi/Siemens)',
        '자재가 위치한 고유 장소 수'
    ])
    
    # Flow Code 분포
    if 'Flow Code' in df.columns:
        flow_counts = df['Flow Code'].value_counts().sort_index()
        for flow_code, count in flow_counts.items():
            flow_desc = {
                0: 'Pre Arrival', 1: 'Port → Site', 2: 'Port → WH → Site',
                3: 'Port → WH → MOSB → Site', 4: 'Multi-hop'
            }.get(flow_code, f'Flow {flow_code}')
            
            summary_data['Metric'].append(f'Flow {flow_code} Count')
            summary_data['Value'].append(f"{count:,}")
            summary_data['Description'].append(f'{flow_desc} 경로를 통한 자재 수량')
    
    # 위치별 분포
    if 'Final Location' in df.columns:
        location_counts = df['Final Location'].value_counts().head(10)
        for location, count in location_counts.items():
            summary_data['Metric'].append(f'{location} Count')
            summary_data['Value'].append(f"{count:,}")
            summary_data['Description'].append(f'{location}에 위치한 자재 수량')
    
    return pd.DataFrame(summary_data)

def style_excel_worksheet(worksheet, df, sheet_name):
    """Excel 워크시트 스타일링"""
    print(f"🎨 {sheet_name} 시트 스타일링 중...")
    
    # 헤더 스타일
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 데이터 스타일
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(
        left=Side(style='thin', color='CCCCCC'), 
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), 
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # 헤더 행 스타일 적용
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # 데이터 행 스타일 적용
    for row_num in range(2, len(df) + 2):
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.alignment = data_alignment
            cell.border = data_border
            
            # 짝수 행 배경색
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    # 컬럼 너비 자동 조정
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # 행 높이 설정
    worksheet.row_dimensions[1].height = 30  # 헤더 행
    for row_num in range(2, len(df) + 2):
        worksheet.row_dimensions[row_num].height = 20  # 데이터 행
    
    # 조건부 서식 (데이터 품질 점수)
    if 'Data Quality Score' in df.columns:
        quality_col = None
        for idx, col in enumerate(df.columns, 1):
            if col == 'Data Quality Score':
                quality_col = idx
                break
        
        if quality_col:
            # 데이터 품질 점수 컬럼에 컬러 스케일 적용
            color_scale = ColorScaleRule(
                start_type='num', start_value=0, start_color='FF6B6B',
                mid_type='num', mid_value=70, mid_color='FFE66D', 
                end_type='num', end_value=100, end_color='4ECDC4'
            )
            
            quality_range = f"{openpyxl.utils.get_column_letter(quality_col)}2:{openpyxl.utils.get_column_letter(quality_col)}{len(df)+1}"
            worksheet.conditional_formatting.add(quality_range, color_scale)

def create_sku_detail_excel():
    """Case No.별 상세 Excel 파일 생성"""
    print("🚀 HVDC SKU Master Hub - Case No.별 상세 Excel 생성 시작")
    print("=" * 60)
    
    try:
        # 1. 데이터 로드
        df = load_sku_master_data()
        
        # 2. 데이터 전처리
        df_enhanced = enhance_dataframe(df)
        
        # 3. 요약 통계 생성
        summary_df = create_summary_stats(df_enhanced)
        
        # 4. Excel 파일 생성
        output_file = f"out/HVDC_SKU_Master_Detail_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        print(f"📝 Excel 파일 생성 중: {output_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 📊 Dashboard 시트 (요약)
            summary_df.to_excel(writer, sheet_name='📊 Dashboard', index=False)
            
            # 📋 SKU Master Detail 시트 (전체 상세 데이터)
            df_enhanced.to_excel(writer, sheet_name='📋 SKU Master Detail', index=False)
            
            # 🎯 By Location 시트 (위치별 정렬)
            if 'Final Location' in df_enhanced.columns:
                df_by_location = df_enhanced.sort_values(['Final Location', 'SKU (Case No.)'])
                df_by_location.to_excel(writer, sheet_name='🎯 By Location', index=False)
            
            # 🔄 By Flow Code 시트 (Flow별 정렬)  
            if 'Flow Code' in df_enhanced.columns:
                df_by_flow = df_enhanced.sort_values(['Flow Code', 'SKU (Case No.)'])
                df_by_flow.to_excel(writer, sheet_name='🔄 By Flow Code', index=False)
            
            # 🏭 By Vendor 시트 (벤더별 정렬)
            if 'Vendor' in df_enhanced.columns:
                df_by_vendor = df_enhanced.sort_values(['Vendor', 'SKU (Case No.)'])
                df_by_vendor.to_excel(writer, sheet_name='🏭 By Vendor', index=False)
            
            # 📦 Heavy Items 시트 (중량 상위 10%)
            if 'Gross Weight (kg)' in df_enhanced.columns:
                threshold = df_enhanced['Gross Weight (kg)'].quantile(0.9)
                df_heavy = df_enhanced[df_enhanced['Gross Weight (kg)'] >= threshold].sort_values('Gross Weight (kg)', ascending=False)
                if not df_heavy.empty:
                    df_heavy.to_excel(writer, sheet_name='📦 Heavy Items', index=False)
        
        # 5. 스타일링 적용
        print("🎨 Excel 스타일링 적용 중...")
        workbook = openpyxl.load_workbook(output_file)
        
        # 각 시트에 스타일 적용
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            if sheet_name == '📊 Dashboard':
                style_excel_worksheet(worksheet, summary_df, sheet_name)
            else:
                style_excel_worksheet(worksheet, df_enhanced, sheet_name)
        
        # 첫 번째 시트를 Dashboard로 설정
        workbook.active = workbook['📊 Dashboard']
        workbook.save(output_file)
        
        # 6. 결과 리포트
        print("\n" + "=" * 60)
        print("✅ HVDC SKU Master Hub - Case No.별 상세 Excel 생성 완료!")
        print("=" * 60)
        print(f"📄 파일명: {output_file}")
        print(f"📊 총 SKU 수: {len(df_enhanced):,}개")
        print(f"📋 컬럼 수: {len(df_enhanced.columns)}개")
        print(f"📑 시트 수: {len(workbook.sheetnames)}개")
        
        # 시트별 정보
        print("\n🗂️ 생성된 시트 목록:")
        for idx, sheet_name in enumerate(workbook.sheetnames, 1):
            if sheet_name == '📊 Dashboard':
                print(f"  {idx}. {sheet_name} - 프로젝트 요약 통계 및 KPI")
            elif 'SKU Master Detail' in sheet_name:
                print(f"  {idx}. {sheet_name} - 전체 {len(df_enhanced):,}개 SKU 상세 정보")
            elif 'By Location' in sheet_name:
                locations = df_enhanced['Final Location'].nunique() if 'Final Location' in df_enhanced.columns else 0
                print(f"  {idx}. {sheet_name} - {locations}개 위치별 SKU 정렬")
            elif 'By Flow Code' in sheet_name:
                flows = df_enhanced['Flow Code'].nunique() if 'Flow Code' in df_enhanced.columns else 0
                print(f"  {idx}. {sheet_name} - {flows}개 Flow별 SKU 정렬")
            elif 'By Vendor' in sheet_name:
                vendors = df_enhanced['Vendor'].nunique() if 'Vendor' in df_enhanced.columns else 0
                print(f"  {idx}. {sheet_name} - {vendors}개 벤더별 SKU 정렬")
            elif 'Heavy Items' in sheet_name:
                heavy_count = len(df_enhanced[df_enhanced['Gross Weight (kg)'] >= df_enhanced['Gross Weight (kg)'].quantile(0.9)]) if 'Gross Weight (kg)' in df_enhanced.columns else 0
                print(f"  {idx}. {sheet_name} - 중량 상위 10% ({heavy_count}개 SKU)")
        
        print(f"\n🎯 사용 방법:")
        print(f"  1. Excel에서 {output_file} 파일 열기")
        print(f"  2. '📊 Dashboard' 시트에서 전체 현황 확인")
        print(f"  3. '📋 SKU Master Detail' 시트에서 Case No.별 상세 정보 조회")
        print(f"  4. 필터/정렬 기능으로 원하는 조건의 SKU 검색")
        print(f"  5. 각 분류별 시트에서 목적에 맞는 데이터 확인")
        
        return output_file
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        raise

if __name__ == "__main__":
    create_sku_detail_excel()
